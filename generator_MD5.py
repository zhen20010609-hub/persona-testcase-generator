# -*- coding: utf-8 -*-
"""
@Author: polly
@Description: MD5测试用例生成器（支持多个扩展字段；扩展字段始终明文输出；正向组合自动生成全部非空子集）
"""

import json
import hashlib
import os
from itertools import combinations
from openpyxl import Workbook


def calculate_md5(value):
    return hashlib.md5(str(value).encode("utf-8")).hexdigest()


def swap_case(text):
    text = str(text)
    return ''.join(char.lower() if char.isupper() else char.upper() for char in text)


def safe_str(value):
    return "" if value is None else str(value).strip()


def modify_string(value, operation):
    value = safe_str(value)
    if operation == "remove_last":
        return value[:-1]
    elif operation == "add_zero":
        return value + "0"
    elif operation == "replace_last_a":
        return value[:-1] + "a" if value else "a"
    return value


def get_positive_key_combinations(extra_fields=None):
    extra_fields = extra_fields or {}
    all_keys = ["id", "cell", "name"] + list(extra_fields.keys())

    results = []
    for r in range(len(all_keys), 0, -1):
        for combo in combinations(all_keys, r):
            combo_list = list(combo)
            desc = f"{len(combo_list)}key-{','.join(combo_list)}"
            results.append((combo_list, desc))
    return results


def build_positive_case_data(input_data, keys, extra_fields=None):
    data = {
        "products": input_data["products"],
        "country": input_data["country"]["plain"]
    }

    if "id" in keys:
        data["id"] = input_data["id"]["plain"]
        if input_data.get("id_type"):
            data["id_type"] = input_data["id_type"]

    if "cell" in keys:
        data["cell"] = input_data["cell"]["plain"]

    if "name" in keys:
        data["name"] = input_data["name"]["plain"]

    if extra_fields:
        for extra_key in extra_fields.keys():
            if extra_key in keys:
                data[extra_key] = input_data[extra_key]["plain"]

    return data


def build_base_data(input_data, extra_fields=None):
    data = {
        "products": input_data["products"],
        "country": input_data["country"]["plain"]
    }

    if extra_fields:
        for field in extra_fields:
            if field in input_data:
                data[field] = input_data[field]["plain"]

    return data


def generate_test_case(base_data, config, fields):
    data = base_data.copy()

    for field, value_type in config.get("fields", {}).items():
        if field in fields:
            if value_type == "md5":
                data[field] = fields[field]["md5"]
            elif value_type == "plain":
                data[field] = fields[field]["plain"]
            elif value_type == "swapped":
                data[field] = swap_case(fields[field]["md5"])
            elif value_type == "invalid_md5":
                data[field] = config["invalid_values"][field]
            elif value_type == "modified_md5":
                modified_value = modify_string(fields[field]["plain"], config["modification"])
                data[field] = calculate_md5(modified_value)
            elif value_type == "invalid_length":
                data[field] = fields[field]["md5"][:-1] if config["length"] == "short" else fields[field]["md5"] + "0"
            elif value_type == "invalid_parse":
                data[field] = fields[field]["md5"][:-1] + "a"

    if "id" in config["fields"] and fields.get("id_type"):
        id_type_value = fields["id_type"]
        if id_type_value:
            data["id_type"] = id_type_value

    return data, config["description"]


def generate_test_cases(products, id_value, id_type, cell, name, country, extra_fields=None):
    extra_fields = extra_fields or {}
    extra_fields = {safe_str(k): str(v) if v is not None else "" for k, v in extra_fields.items() if safe_str(k)}

    input_data = {
        "products": safe_str(products),
        "id": {"plain": safe_str(id_value)},
        "cell": {"plain": safe_str(cell)},
        "name": {"plain": safe_str(name)},
        "country": {"plain": safe_str(country)},
        "id_type": safe_str(id_type)
    }

    for extra_key, extra_value in extra_fields.items():
        input_data[extra_key] = {"plain": extra_value}

    for field, field_value in input_data.items():
        if isinstance(field_value, dict) and "plain" in field_value and field != "country":
            field_value["md5"] = calculate_md5(field_value["plain"])

    json_data = []
    descriptions = []

    key_combinations = get_positive_key_combinations(extra_fields)
    for keys, desc in key_combinations:
        data = build_positive_case_data(input_data, keys, extra_fields=extra_fields)
        json_data.append(json.dumps(data, ensure_ascii=False))
        descriptions.append(f"{len(descriptions) + 1}. MD5-Forward_Check-{desc}")

    base_data = build_base_data(input_data, extra_fields=extra_fields)

    invalid_md5_values = {
        "id": {
            "empty": "d41d8cd98f00b204e9800998ecf8427e",
            "space": "7215ee9c7d9dc229d2921a40e899ec5f",
            "null": "37a6259cc0c1dae299a7866489dff0bd",
            "none": "b50339a10e1de285ac99d4c3990b8693",
            "na": "14795de509a66f9b97f5cdebde91fa0c"
        },
        "cell": {
            "empty": "d41d8cd98f00b204e9800998ecf8427e",
            "space": "7215ee9c7d9dc229d2921a40e899ec5f",
            "array": "5f1cb4e5159145f28dc6b9176b2c2ef4",
            "null": "37a6259cc0c1dae299a7866489dff0bd",
            "none": "b50339a10e1de285ac99d4c3990b8693",
            "na": "14795de509a66f9b97f5cdebde91fa0c"
        },
        "name": {
            "empty": "d41d8cd98f00b204e9800998ecf8427e",
            "space": "7215ee9c7d9dc229d2921a40e899ec5f",
            "null": "37a6259cc0c1dae299a7866489dff0bd",
            "none": "b50339a10e1de285ac99d4c3990b8693",
            "na": "14795de509a66f9b97f5cdebde91fa0c",
            "A": "7fc56270e7a70fa81a5935b72eacbe29",
            "123": "202cb962ac59075b964b07152d234b70",
            "chinese": "b458bf3f7ddcc18c2a93bb7e8c1dd482"
        }
    }

    test_configs = [
        {"fields": {"id": "plain", "cell": "plain", "name": "plain"}, "description": "MD5-Forward_Check-Plain-id,cell,name_Plain"},
        {"fields": {"id": "md5", "cell": "plain", "name": "plain"}, "description": "MD5-Forward_Check-MD5+Plain-id_md5, cell、name_plain"},
        {"fields": {"id": "plain", "cell": "md5", "name": "plain"}, "description": "MD5-Forward_Check-MD5+Plain-cell_md5, id、name_plain"},
        {"fields": {"id": "plain", "cell": "plain", "name": "md5"}, "description": "MD5-Forward_Check-MD5+Plain-name_md5, cell、id_plain"},
        {"fields": {"id": "md5", "cell": "md5", "name": "md5"}, "description": "MD5-Forward_Check-only_md5-id,cell,name"},
        {"fields": {"id": "swapped", "cell": "swapped", "name": "md5"}, "description": "MD5-Forward_Check-only_md5-id_Case_Swap"},
    ]

    base_fields = ["id", "cell", "name"]
    for r in range(1, len(base_fields) + 1):
        for combo in combinations(base_fields, r):
            test_configs.append({
                "fields": {field: "md5" for field in combo},
                "description": f"MD5-Forward_Check-only_md5-{','.join(combo)}"
            })

    error_cases = [
        *[
            {
                "fields": {"id": "invalid_md5", "cell": "md5", "name": "md5"},
                "invalid_values": {"id": invalid_md5_values["id"][case]},
                "description": f"MD5-Abnormal_Check-wrong_md5-id-{case.upper()}"
            }
            for case in ["empty", "space", "null", "none", "na"]
        ],
        *[
            {
                "fields": {"id": "modified_md5", "cell": "md5", "name": "md5"},
                "modification": mod,
                "description": f"MD5-Abnormal_Check-id is：{desc}"
            }
            for mod, desc in [
                ("remove_last", "one_less_digit"),
                ("add_zero", "one_more_digit"),
                ("replace_last_a", "format_error")
            ]
        ],
        *[
            {
                "fields": {"id": "md5", "cell": "invalid_md5", "name": "md5"},
                "invalid_values": {"cell": invalid_md5_values["cell"][case]},
                "description": f"MD5-Abnormal_Check-wrong_md5-cell-{case.upper() if case != 'array' else '[]'}"
            }
            for case in ["empty", "space", "array", "null", "none", "na"]
        ],
        *[
            {
                "fields": {"id": "md5", "cell": "modified_md5", "name": "md5"},
                "modification": mod,
                "description": f"MD5-Abnormal_Check-wrong_md5-cell{'one_less_digit' if mod == 'remove_last' else 'one_more_digit' if mod == 'add_zero' else 'format_error'}"
            }
            for mod in ["remove_last", "add_zero", "replace_last_a"]
        ],
        *[
            {
                "fields": {"id": "md5", "cell": "md5", "name": "invalid_md5"},
                "invalid_values": {"name": invalid_md5_values["name"][case]},
                "description": f"MD5-Abnormal_Check-wrong_md5-name-{case.upper() if case not in ['chinese', '123'] else 'áéíóúüÁÉÍÓÚÜñÑ' if case == 'chinese' else '123'}"
            }
            for case in ["empty", "space", "null", "none", "na", "A", "123", "chinese"]
        ],
        *[
            {
                "fields": {field: "invalid_length", **{f: "md5" for f in base_fields if f != field}},
                "length": length,
                "description": f"MD5-Abnormal_Check-invalid_md5-length-{field}-{31 if length == 'short' else 33}-bit"
            }
            for field in base_fields for length in ["short", "long"]
        ],
        *[
            {
                "fields": {field: "invalid_parse", **{f: "md5" for f in base_fields if f != field}},
                "description": f"MD5-Abnormal_Check-invalid_md5-decryption_failed-{field}32-bit_decryption_failed"
            }
            for field in base_fields
        ]
    ]

    test_configs.extend(error_cases)

    for config in test_configs:
        data, desc = generate_test_case(base_data, config, input_data)
        json_data.append(json.dumps(data, ensure_ascii=False))
        descriptions.append(f"{len(descriptions) + 1}. {desc}")

    for extra_key, extra_value in extra_fields.items():
        for raw_value, desc_suffix in [("", "isEmpty"), (" ", "isSpace")]:
            data = build_base_data(input_data, extra_fields=extra_fields)
            data["id"] = input_data["id"]["md5"]
            data["cell"] = input_data["cell"]["md5"]
            data["name"] = input_data["name"]["md5"]
            if input_data.get("id_type"):
                data["id_type"] = input_data["id_type"]
            data[extra_key] = raw_value

            json_data.append(json.dumps(data, ensure_ascii=False))
            descriptions.append(f"{len(descriptions) + 1}. MD5-Abnormal_Check-{extra_key}{desc_suffix}")

        for operation, desc_suffix in [
            ("remove_last", "One less digit"),
            ("add_zero", "One more digit"),
            ("replace_last_a", "Format error")
        ]:
            data = build_base_data(input_data, extra_fields=extra_fields)
            data["id"] = input_data["id"]["md5"]
            data["cell"] = input_data["cell"]["md5"]
            data["name"] = input_data["name"]["md5"]
            if input_data.get("id_type"):
                data["id_type"] = input_data["id_type"]

            data[extra_key] = modify_string(extra_value, operation)

            json_data.append(json.dumps(data, ensure_ascii=False))
            descriptions.append(f"{len(descriptions) + 1}. MD5-Abnormal_Check-{extra_key}{desc_suffix}")

    return json_data, descriptions


def save_to_excel(json_data, descriptions, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "testcases"

    ws["A1"] = "num"
    ws["B1"] = "Description"
    ws["C1"] = "req_data"

    for i, (desc, json_str) in enumerate(zip(descriptions, json_data), start=1):
        ws.cell(row=i + 1, column=1, value=i)
        ws.cell(row=i + 1, column=2, value=desc)
        ws.cell(row=i + 1, column=3, value=json_str)

    wb.save(filename)


def save_metadata_excel_without_req_data(descriptions, output_excel_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "testcases"

    ws["A1"] = "num"
    ws["B1"] = "Description"
    ws["C1"] = "req_data"

    for i, desc in enumerate(descriptions, start=1):
        ws.cell(row=i + 1, column=1, value=i)
        ws.cell(row=i + 1, column=2, value=desc)
        ws.cell(row=i + 1, column=3, value="")

    wb.save(output_excel_path)


def save_to_txt_files(json_data, output_dir):
    os.makedirs(output_dir, exist_ok=True)

    for i, json_str in enumerate(json_data, start=1):
        file_path = os.path.join(output_dir, f"{i}.txt")
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(json_str)


def generate_md5_test_cases_excel(products, id_value, id_type, cell, name, country, output_path, extra_fields=None):
    results, descs = generate_test_cases(
        products=products,
        id_value=id_value,
        id_type=id_type,
        cell=cell,
        name=name,
        country=country,
        extra_fields=extra_fields
    )
    save_to_excel(results, descs, output_path)
    return output_path


def generate_md5_test_cases_txt(products, id_value, id_type, cell, name, country, output_dir, extra_fields=None):
    results, descs = generate_test_cases(
        products=products,
        id_value=id_value,
        id_type=id_type,
        cell=cell,
        name=name,
        country=country,
        extra_fields=extra_fields
    )
    save_to_txt_files(results, output_dir)
    save_metadata_excel_without_req_data(descs, os.path.join(output_dir, "bigFile_output-MD5.xlsx"))
    return output_dir


if __name__ == "__main__":
    output_file = "1-AltScoreTelco_PH-weakVerify-md5.xlsx"
    generate_md5_test_cases_excel(
        products="AltScoreTelco_PH",
        id_value="011115634849",
        id_type="UMID",
        cell="09206587342",
        name="aaa",
        country="PH",
        output_path=output_file,
        extra_fields={"gaid": "A1B2C3D4"}
    )
    print(f"Data has been successfully exported: {output_file}")