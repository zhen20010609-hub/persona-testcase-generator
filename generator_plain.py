# -*- coding: utf-8 -*-
"""
@Author: polly
@Description: 明文测试用例生成器（支持多个扩展字段；正向组合自动生成全部非空子集）
"""

import json
import hashlib
import os
from itertools import combinations
from openpyxl import Workbook


def generate_md5(value):
    return hashlib.md5(str(value).encode("utf-8")).hexdigest()


def swap_case(value):
    value = str(value)
    return ''.join(char.lower() if char.isupper() else char.upper() for char in value)


def safe_str(value):
    return "" if value is None else str(value).strip()


def build_base_data(inputs, extra_fields=None):
    data = {
        "products": inputs["products"],
        "id": inputs["id"],
        "cell": inputs["cell"],
        "name": inputs["name"],
        "country": inputs["country"]
    }

    if inputs.get("id_type"):
        data["id_type"] = inputs["id_type"]

    if extra_fields:
        data.update(extra_fields)

    return data


def get_invalid_modifications(value, is_id_country_special=False):
    value = safe_str(value)

    if is_id_country_special:
        return [
            (value[:9], "One less digit"),
            (value.ljust(14, '0'), "One more digit"),
            ((value[:-1] + "a") if value else "a", "Format error")
        ]

    return [
        (value[:-1], "One less digit"),
        (value + "0", "One more digit"),
        ((value[:-1] + "a") if value else "a", "Format error")
    ]


def get_positive_key_combinations(extra_fields=None):
    extra_fields = extra_fields or {}
    all_keys = ["id", "cell", "name"] + list(extra_fields.keys())

    results = []
    for r in range(len(all_keys), 0, -1):
        for combo in combinations(all_keys, r):
            combo_list = list(combo)
            desc = f"{len(combo_list)}key-{','.join(combo_list)}"
            results.append((combo_list, desc))
    return results


def generate_test_cases(products, id_value, id_type, cell, name, country, extra_fields=None):
    extra_fields = extra_fields or {}
    extra_fields = {safe_str(k): str(v) if v is not None else "" for k, v in extra_fields.items() if safe_str(k)}

    inputs = {
        "products": safe_str(products),
        "id": safe_str(id_value),
        "id_type": safe_str(id_type),
        "cell": safe_str(cell),
        "name": safe_str(name),
        "country": safe_str(country)
    }

    key_combinations = get_positive_key_combinations(extra_fields)

    invalid_values = [
        ("", "isEmpty"),
        (" ", "isSpace"),
        ("null", "isnull"),
        ("NONE", "isNONE"),
        ("NA/N", "isNA/N")
    ]

    invalid_name_cases = [
        ("A", "is A"),
        ("abc", "is abc"),
        ("123", "is 123")
    ]

    country_cases = [
        ("AB", "country:AB"),
        ("", "country isEmpty"),
        (" ", "country isSpace"),
        (None, "country notTrans"),
        ("CO", "country isCO"),
    ]

    json_data = []
    descriptions = []
    case_number = 1

    def append_case(data, desc):
        nonlocal case_number
        json_data.append(json.dumps(data, ensure_ascii=False))
        descriptions.append(f"{case_number}. {desc}")
        case_number += 1

    for keys, desc in key_combinations:
        data = {
            "products": inputs["products"],
            "country": inputs["country"]
        }

        if "id" in keys:
            data["id"] = inputs["id"]
            if inputs.get("id_type"):
                data["id_type"] = inputs["id_type"]

        if "cell" in keys:
            data["cell"] = inputs["cell"]

        if "name" in keys:
            data["name"] = inputs["name"]

        for extra_key, extra_value in extra_fields.items():
            if extra_key in keys:
                data[extra_key] = extra_value

        append_case(data, f"Plaintext-Forward_Check-{desc}")

    special_cases = [
        ("id_case_swap", "Plain-Forward_Check-Input_Validation-id_Case_Swap"),
        ("cell_multi", "Plain-Forward_Check-Input_Validation-cell_Multi-value_Input"),
        ("name_special", "Plain-Forward_Check-Input_Validation-name_isHERNÁNDEZ MORENO BRENDA ALEJANDRA"),
        ("id_md5", "Plain-Forward_Check-MD5+Plain-id_md5，cell、name_Plain"),
        ("cell_md5", "Plain-Forward_Check-MD5+Plain-cell_md5，id、name_Plain"),
        ("name_md5", "Plain-Forward_Check-MD5+Plain-name_md5，cell、id_Plain"),
        ("all_md5", "Plain-Forward_Check-only-MD5-id、cell、name md5")
    ]

    for transform_key, desc in special_cases:
        data = build_base_data(inputs, extra_fields=extra_fields)

        if transform_key == "id_case_swap":
            data["id"] = swap_case(inputs["id"])
        elif transform_key == "cell_multi":
            data["cell"] = [
                inputs["cell"],
                inputs["cell"][:-6] + "123123" if len(inputs["cell"]) >= 6 else "123123"
            ]
        elif transform_key == "name_special":
            data["name"] = "HERNÁNDEZ MORENO BRENDA ALEJANDRA"
        elif transform_key == "id_md5":
            data["id"] = generate_md5(inputs["id"])
        elif transform_key == "cell_md5":
            data["cell"] = generate_md5(inputs["cell"])
        elif transform_key == "name_md5":
            data["name"] = generate_md5(inputs["name"])
        elif transform_key == "all_md5":
            data["id"] = generate_md5(inputs["id"])
            data["cell"] = generate_md5(inputs["cell"])
            data["name"] = generate_md5(inputs["name"])

        append_case(data, desc)

    for key in ["id", "cell", "name"]:
        for value, desc in invalid_values:
            data = build_base_data(inputs, extra_fields=extra_fields)
            data[key] = value
            append_case(data, f"Plain-Abnormal_Check-{key}{desc}")

    for key in ["id", "cell"]:
        is_special = (inputs["country"] == "ID" and key == "cell")
        modifications = get_invalid_modifications(inputs[key], is_id_country_special=is_special)

        for mod_value, desc in modifications:
            data = build_base_data(inputs, extra_fields=extra_fields)
            data[key] = mod_value
            append_case(data, f"Plain-Abnormal_Check-{key}{desc}")

    for value, desc in invalid_name_cases:
        data = build_base_data(inputs, extra_fields=extra_fields)
        data["name"] = value
        append_case(data, f"Plain-Abnormal_Check-name{desc}")

    for value, desc in country_cases:
        data = {
            "products": inputs["products"],
            "id": inputs["id"],
            "cell": inputs["cell"],
            "name": inputs["name"]
        }

        if value is not None:
            data["country"] = value

        if inputs.get("id_type"):
            data["id_type"] = inputs["id_type"]

        if extra_fields:
            data.update(extra_fields)

        append_case(data, f"Plain-Abnormal_Check-countryCheck-{desc}")

    if inputs["country"] == "PH":
        ph_special_cases = [
            {"products": inputs["products"], "cell": "09951822873", "id": "33597012480", "name": "Anthony", "id_type": "SSS", "country": inputs["country"]},
            {"products": inputs["products"], "cell": "09951822873", "id": "00455090", "name": "Anthony", "id_type": "PRC", "country": inputs["country"]},
            {"products": inputs["products"], "cell": "09951822873", "id": "A12345678900", "name": "Anthony", "id_type": "DLN", "country": inputs["country"]},
            {"products": inputs["products"], "cell": "09951822873", "id": "P12345670", "name": "Anthony", "id_type": "PPN", "country": inputs["country"]},
            {"products": inputs["products"], "cell": "09951822873", "id": "1212345678910", "name": "Anthony", "id_type": "PHN", "country": inputs["country"]},
            {"products": inputs["products"], "cell": "09951822873", "id": "D012345678900", "name": "Anthony", "id_type": "PCN", "country": inputs["country"]},
            {"products": inputs["products"], "cell": "09951822873", "id": "A6011099224", "name": "Anthony", "id_type": "GSIS", "country": inputs["country"]},
            {"products": inputs["products"], "cell": "09951822873", "id": "A23456789", "name": "Anthony", "id_type": "TIN", "country": inputs["country"]},
            {"products": inputs["products"], "cell": "09951822873", "id": "1234567890123456A", "name": "Anthony", "id_type": "PSN", "country": inputs["country"]},
            {"products": inputs["products"], "cell": "09951822873", "id": "1234567890123456A", "name": "Anthony", "id_type": "", "country": inputs["country"]}
        ]

        for data in ph_special_cases:
            if extra_fields:
                data.update(extra_fields)
            append_case(data, f"Plain-Forward_Check-PH_idTpye-{data['id_type']}")

    for extra_key, extra_value in extra_fields.items():
        for abnormal_value, abnormal_desc in [("", "isEmpty"), (" ", "isSpace")]:
            data = build_base_data(inputs, extra_fields=extra_fields)
            data[extra_key] = abnormal_value
            append_case(data, f"Plain-Abnormal_Check-{extra_key}{abnormal_desc}")

        for abnormal_value, abnormal_desc in get_invalid_modifications(extra_value):
            data = build_base_data(inputs, extra_fields=extra_fields)
            data[extra_key] = abnormal_value
            append_case(data, f"Plain-Abnormal_Check-{extra_key}{abnormal_desc}")

    return json_data, descriptions


def save_to_excel(json_data, descriptions, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "testcases"

    ws["A1"] = "num"
    ws["B1"] = "Description"
    ws["C1"] = "req_data"

    for i, (desc, json_str) in enumerate(zip(descriptions, json_data), start=1):
        ws.cell(row=i + 1, column=1, value=i)
        ws.cell(row=i + 1, column=2, value=desc)
        ws.cell(row=i + 1, column=3, value=json_str)

    wb.save(filename)


def save_metadata_excel_without_req_data(descriptions, output_excel_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "testcases"

    ws["A1"] = "num"
    ws["B1"] = "Description"
    ws["C1"] = "req_data"

    for i, desc in enumerate(descriptions, start=1):
        ws.cell(row=i + 1, column=1, value=i)
        ws.cell(row=i + 1, column=2, value=desc)
        ws.cell(row=i + 1, column=3, value="")

    wb.save(output_excel_path)


def save_to_txt_files(json_data, output_dir):
    os.makedirs(output_dir, exist_ok=True)

    for i, json_str in enumerate(json_data, start=1):
        file_path = os.path.join(output_dir, f"{i}.txt")
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(json_str)


def generate_plain_test_cases_excel(products, id_value, id_type, cell, name, country, output_path, extra_fields=None):
    results, descs = generate_test_cases(
        products=products,
        id_value=id_value,
        id_type=id_type,
        cell=cell,
        name=name,
        country=country,
        extra_fields=extra_fields
    )
    save_to_excel(results, descs, output_path)
    return output_path


def generate_plain_test_cases_txt(products, id_value, id_type, cell, name, country, output_dir, extra_fields=None):
    results, descs = generate_test_cases(
        products=products,
        id_value=id_value,
        id_type=id_type,
        cell=cell,
        name=name,
        country=country,
        extra_fields=extra_fields
    )
    save_to_txt_files(results, output_dir)
    save_metadata_excel_without_req_data(descs, os.path.join(output_dir, "bigFile_output-plain.xlsx"))
    return output_dir


if __name__ == "__main__":
    output_file = "1-AltScoreTelco_PH-weakVerify-plain.xlsx"
    generate_plain_test_cases_excel(
        products="AltScoreTelco_PH",
        id_value="011115634849",
        id_type="UMID",
        cell="09206587342",
        name="aaa",
        country="PH",
        output_path=output_file,
        extra_fields={"gaid": "A1B2C3D4"}
    )
    print(f"Data has been successfully exported: {output_file}")